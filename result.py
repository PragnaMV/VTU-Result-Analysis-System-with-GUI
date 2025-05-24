import pdfplumber
import re
import os
import pymysql.cursors
import pandas as pd  # Import pandas library for handling dataframes
import zipfile
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import warnings
from openpyxl.styles import Border, Side
from pathlib import Path
import sys

import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

def send_email_with_attachment(recipient, subject, body, attachment_path):
    # Temporarily hardcode email credentials for testing
    sender_email = ""  # Replace with your email
    sender_password = "" # Replace with your app password

    try:
        smtp_server = "smtp.gmail.com"
        smtp_port = 587

        # Create the email message
        msg = MIMEMultipart()
        msg["From"] = sender_email
        msg["To"] = recipient
        msg["Subject"] = subject
        msg.attach(MIMEText(body, "plain"))

        # Attach the file
        with open(attachment_path, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                "Content-Disposition",
                f"attachment; filename={attachment_path.split('/')[-1]}",
            )
            msg.attach(part)

        # Connect to the SMTP server and send the email
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender_email, sender_password)
            server.sendmail(sender_email, recipient, msg.as_string())

        print(f"Email sent successfully to {recipient}")
    except Exception as e:
        print(f"Failed to send email to {recipient}. Error: {str(e)}")

def notify_faculty(attachment_path):
    # List of faculty email addresses
    faculty_emails = [
        
        ""
    ]
    subject = "Student Results Report"
    body = "Dear Faculty,\n\nPlease find attached the latest student results report.\n\nRegards,\nResult Analysis Team"

    # Send the email to each faculty member
    for email in faculty_emails:
        send_email_with_attachment(email, subject, body, attachment_path)


# Connect to MySQL database
connection = pymysql.connect(
    host='localhost',
    user='root',
    password='', #your password
    database='', #your database
    cursorclass=pymysql.cursors.DictCursor
)

def create_tables():
    with connection.cursor() as cursor:
        # Create table for extracted data
        sql_create_extracted_data_table = """
        CREATE TABLE IF NOT EXISTS extracted_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SubjectCode VARCHAR(255),
            SubjectName VARCHAR(255),
            InternalMarks INT,
            ExternalMarks INT,
            TotalMarks INT,
            Result VARCHAR(255),
            AnnouncedDate DATE,
            INDEX idx_usn (Usn),
            INDEX idx_semester (Semester),
            INDEX idx_student_name (StudentName),
            INDEX idx_subject_code (SubjectCode),
            INDEX idx_subject_name (SubjectName)
        )
        """
        cursor.execute(sql_create_extracted_data_table)

        
        # Create table for SGPA data
        sql_create_sgpa_table = """
        CREATE TABLE IF NOT EXISTS sgpa_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SGPA FLOAT,
            INDEX idx_sgpa (SGPA),  -- Add index for the SGPA column
            FOREIGN KEY (Usn) REFERENCES extracted_data(Usn),
            FOREIGN KEY (Semester) REFERENCES extracted_data(Semester),
            FOREIGN KEY (StudentName) REFERENCES extracted_data(StudentName)
        )

        """
        cursor.execute(sql_create_sgpa_table)
        # Create table for storing top performers
        sql_create_topper_table = """
        CREATE TABLE IF NOT EXISTS topper_data (
            id INT AUTO_INCREMENT PRIMARY KEY,
            Usn VARCHAR(255),
            Semester VARCHAR(255),
            StudentName VARCHAR(255),
            SGPA FLOAT,
            INDEX idx_usn (Usn),
            INDEX idx_semester (Semester),
            INDEX idx_student_name (StudentName),
            FOREIGN KEY (Usn) REFERENCES extracted_data(Usn),
            FOREIGN KEY (Semester) REFERENCES extracted_data(Semester),
            FOREIGN KEY (StudentName) REFERENCES extracted_data(StudentName),
            FOREIGN KEY (SGPA) REFERENCES sgpa_data(SGPA)
        )
        """
        cursor.execute(sql_create_topper_table)

def insert_topper_data(data):
    with connection.cursor() as cursor:
        sql_insert_topper = """
        INSERT INTO topper_data (Usn, Semester, StudentName, SGPA) 
        VALUES (%s, %s, %s, %s)
        """
        for index, row in data.iterrows():
            cursor.execute(sql_insert_topper, (row['University Seat Number'], row['Semester'], row['Student Name'], row['SGPA']))
    connection.commit()

def insert_extracted_data(data):
    with connection.cursor() as cursor:
        sql_insert_data = """
        INSERT INTO extracted_data 
        (Usn, Semester, StudentName, SubjectCode, SubjectName, InternalMarks, ExternalMarks, TotalMarks, Result, AnnouncedDate) 
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        for row in data:
            # Handle result 'W' case (Withheld)
            if row['Result'] == "W":
                row['Total Marks'] = row['Internal Marks']  # Set total marks equal to internal if withheld
            cursor.execute(sql_insert_data, (row['Usn'], row['Semester'], row['StudentName'], row['Subject Code'], row['Subject Name'], row['Internal Marks'], row['External Marks'], row['Total Marks'], row['Result'], row['AnnouncedDate']))
    connection.commit()


def insert_subject_details(subject_details):
    with connection.cursor() as cursor:
        sql_insert_details = """
        INSERT INTO subject_details 
        (SubjectCode, SubjectName, Credits) 
        VALUES (%s, %s, %s)
        """
        for subject_code, details in subject_details.items():
            cursor.execute(sql_insert_details, (subject_code, details['SubjectName'], details['Credits']))
    connection.commit()

def fetch_credits_from_scheme(subject_code):
    try:
        with connection.cursor() as cursor:
            print(f"Fetching credits for subject code: {subject_code}")  # Debugging print
            sql_fetch_credits = """
            SELECT credits FROM scheme WHERE code = %s
            """
            cursor.execute(sql_fetch_credits, (subject_code,))
            result = cursor.fetchone()
            if result:
                return result['credits']
            else:
                print(f"Subject code '{subject_code}' not found in the scheme table.")  # Debugging print
                return None  # Return None if subject code not found
    except Exception as e:
        print(f"Error fetching credits from scheme table: {e}")
        return None  # Handle the error properly

def fetch_credits(subject_code):
    if subject_code not in subject_details:
        credits = fetch_credits_from_scheme(subject_code)
        if credits is None:
            credits = 0  # Default to 0 if credits not found
        subject_details[subject_code] = credits
        print(f"Credits for subject '{subject_code}': {credits}")  # Debugging print
    else:
        credits = subject_details[subject_code]
    return credits
        
def calculate_grade_point(total_marks, credits):
    if 90 <= total_marks <= 100:
        return 10
    elif 80 <= total_marks <= 89:
        return 9
    elif 70 <= total_marks <= 79:
        return 8
    elif 60 <= total_marks <= 69:
        return 7
    elif 55 <= total_marks <= 59:
        return 6
    elif 50 <= total_marks <= 54:
        return 5
    elif 40 <= total_marks <= 49:
        return 4
    else:
        return 0

def extract_text_from_pdf(pdf_file_path):
    pdf_document = pdfplumber.open(pdf_file_path)
    extracted_text = ""
    # Loop through the pages using pdf_document.pages, which returns a list of pages
    for page in pdf_document.pages:
        text = page.extract_text()  # Use extract_text instead of get_text
        extracted_text += text
        c = extracted_text.replace(',', '').replace('&', '')
        d = c.replace('/', '')
    return d

def extract_specific_sections(extracted_text):
    patterns = {
        "University Seat Number": r'University\s+Seat\s+Number\s*:\s*(\S+)',
        "Semester": r'Semester\s*:\s*(\S+)',
        "Student Name": r'Student\s+Name\s*:\s*(.+)',
        "Subject Code": r'Subject\s+Code\s*:\s*(\S+)',
        "Subject Name": r'Subject\s+Name\s*:\s*(.+)',
        "Internal Marks": r'Internal\s+Marks\s*:\s*(\d+)',
        "External Marks": r'External\s+Marks\s*:\s*(\d+)',
        "Total Marks": r'Total\s+Marks\s*:\s*(\d+)',
        "Result": r'Result\s*:\s*(\S+)'
    }

    extracted_sections = {}
    for section, pattern in patterns.items():
        match = re.search(pattern, extracted_text, re.IGNORECASE)
        if match:
            extracted_sections[section] = match.group(1).strip()
        else:
            if section == "Semester":
                extracted_sections[section] = ""
            else:
                extracted_sections[section] = "Unknown"
    return extracted_sections

def extract_table_data(extracted_text, extracted_sections):
    # Adjusted regex to allow for optional external marks (empty or zero) and more flexible subject names
    pattern = r'([A-Za-z]+\d+[A-Za-z]*)\s+([\w\s#-]+?)\s+(\d+)\s+(\d*)\s+(\d+)\s+([A-ZW])\s+(\d{4}-\d{2}-\d{2})'
    matches = re.findall(pattern, extracted_text)
    extracted_data = []
    
    for match in matches:
        subject_code, subject_name, internal_marks, external_marks, total_marks, result, announced_date = match
        
        # Handle external marks if missing or empty (replace with 0)
        external_marks = int(external_marks) if external_marks.strip() else 0
        
        # Handle withheld (W) result case
        if result == 'W':  # When result is Withheld, set external marks to 0
            external_marks = 0  # Set external marks to 0 for Withheld results
            total_marks = int(internal_marks)  # Set total marks to internal marks in case of Withheld
        else:
            # If total marks field is missing, set it as sum of internal and external marks
            total_marks = int(total_marks) if total_marks.strip() else int(internal_marks) + external_marks
        
        # Add the data to extracted_data list
        extracted_data.append({
            "Usn": extracted_sections.get("University Seat Number", ""),
            "Semester": extracted_sections.get("Semester", ""),
            "StudentName": extracted_sections.get("Student Name", ""),
            "Subject Code": subject_code.strip(),
            "Subject Name": subject_name.strip(),
            "Internal Marks": int(internal_marks),
            "External Marks": external_marks,
            "Total Marks": total_marks,
            "Result": result.strip(),
            "AnnouncedDate": announced_date.strip()
        })
    return extracted_data
    
def insert_sgpa_data(Usn, Semester, StudentName, SGPA):
    with connection.cursor() as cursor:
        sql_insert_sgpa = """
        INSERT INTO sgpa_data (Usn, Semester, StudentName, SGPA) 
        VALUES (%s, %s, %s, %s)
        """
        cursor.execute(sql_insert_sgpa, (Usn, Semester, StudentName, SGPA))
    connection.commit()

def swap_columns(df, col1, col2):
    col_order = df.columns.tolist()
    try:
        idx1, idx2 = col_order.index(col1), col_order.index(col2)
        col_order[idx1], col_order[idx2] = col_order[idx2], col_order[idx1]
        return df[col_order]
    except ValueError:
        print(f"Column '{col1}' or '{col2}' not found in DataFrame.")
        return df
    
def get_top_toppers(df):
    """Get the top 3 toppers based on SGPA, handling ties."""
    # Rank students based on SGPA with ties
    df['Rank'] = df['SGPA'].rank(method='dense', ascending=False)
    # Select the top 3 ranks, including ties
    toppers_df = df[df['Rank'] <= 3].sort_values(by=['Rank', 'SGPA', 'Student Name'])
    return toppers_df[['Semester', 'University Seat Number', 'Student Name', 'SGPA', 'Rank']]

def insert_topper_data(data):
    with connection.cursor() as cursor:
        sql_insert_topper = """
        INSERT INTO topper_data (Usn, Semester, StudentName, SGPA) 
        VALUES (%s, %s, %s, %s)
        """
        for index, row in data.iterrows():
            cursor.execute(sql_insert_topper, (row['University Seat Number'], row['Semester'], row['Student Name'], row['SGPA']))
    connection.commit()

def calculate_result_counts(data):
    """Calculate the counts of FCD, FC, SC, Fail, and W (Withheld) for each subject."""
    subject_results = defaultdict(lambda: Counter({'FCD': 0, 'FC': 0, 'SC': 0, 'Fail': 0, 'W': 0}))  # Add 'W' for Withheld
    
    for row in data:
        subject_code = row['Subject Code']
        total_marks = row['Total Marks']
        result = row.get('Result', '')
        
        if result == 'F':
            subject_results[subject_code]['Fail'] += 1
        elif result == 'W':  # Handle 'Withheld' result
            subject_results[subject_code]['W'] += 1
        elif total_marks >= 70:
            subject_results[subject_code]['FCD'] += 1
        elif total_marks >= 60:
            subject_results[subject_code]['FC'] += 1
        elif total_marks >= 35:
            subject_results[subject_code]['SC'] += 1
        else:
            subject_results[subject_code]['Fail'] += 1

    return subject_results

def create_result_counts_df(subject_results, total_counts):
    """Convert the subject results count into a DataFrame for Excel export."""
    rows = []
    for subject_code, counts in subject_results.items():
        total = total_counts[subject_code]
        rows.append({
            'Subject Code': subject_code,
            'FCD': f"{counts['FCD']}/{total}",
            'FC': f"{counts['FC']}/{total}",
            'SC': f"{counts['SC']}/{total}",
            'Fail': f"{counts['Fail']}/{total}"
        })
    return pd.DataFrame(rows)

def create_result_counts_df(subject_results, total_counts):
    """Convert the subject results count into a DataFrame for Excel export."""
    rows = []
    for subject_code, counts in subject_results.items():
        total = total_counts[subject_code]
        rows.append({
            'Subject Code': subject_code,
            'FCD': f"{counts['FCD']}/{total}",
            'FC': f"{counts['FC']}/{total}",
            'SC': f"{counts['SC']}/{total}",
            'Fail': f"{counts['Fail']}/{total}",
            'Withheld': f"{counts['W']}/{total}"  # Include "Withheld" results
        })
    return pd.DataFrame(rows)

def create_result_counts_df(subject_results, total_counts):
    """Convert the subject results count into a DataFrame for Excel export."""
    rows = []
    for subject_code, counts in subject_results.items():
        total = total_counts[subject_code]
        rows.append({
            'Subject Code': subject_code,
            'FCD': f"{counts['FCD']}/{total}",
            'FC': f"{counts['FC']}/{total}",
            'SC': f"{counts['SC']}/{total}",
            'Fail': f"{counts['Fail']}/{total}",
            'Withheld': f"{counts['W']}/{total}"  # Include "Withheld" results
        })
    return pd.DataFrame(rows)
    
# Function to clean and truncate the subject name
def clean_and_truncate_subject_name(subject_name, max_length=100):
    # Remove leading and trailing spaces
    subject_name = subject_name.strip()
    # Replace newline and carriage return with a space
    subject_name = subject_name.replace('\n', ' ').replace('\r', ' ')
    # Remove any non-alphanumeric characters except spaces
    subject_name = re.sub(r'[^A-Za-z0-9\s]', '', subject_name)
    # Replace multiple spaces with a single space
    subject_name = re.sub(r'\s+', ' ', subject_name)
    # Truncate the subject name to fit the database column length
    if len(subject_name) > max_length:
        subject_name = subject_name[:max_length]
    return subject_name

def fetch_all_students_from_db():
    """
    Fetch the total number of students from the database.
    """
    try:
        query = "SELECT COUNT(*) AS count FROM student"
        with connection.cursor() as cursor:
            cursor.execute(query)
            db_student_count = cursor.fetchone()['count']
        return db_student_count
    except Exception as e:
        print(f"Error fetching student count from DB: {e}")
        return 0


def validate_student_counts(usn_list):
    """
    Validates that the number of extracted USNs matches the number of student records in the database.
    """
    try:
        # Count of extracted students
        extracted_student_count = len(usn_list)

        # Fetch count from the database
        db_student_count = fetch_all_students_from_db()

        # Validate and report
        if db_student_count == extracted_student_count:
            print(f"Validation successful: {db_student_count} students in DB match {extracted_student_count} students extracted.")
        else:
            print(f"Validation failed: {db_student_count} students in DB do not match {extracted_student_count} students extracted.")
            print("Please check for discrepancies in the data.")
    except Exception as e:
        print(f"Error during validation: {e}")

def validate_student_counts(connection, excel_file_path):
    """
    Validates that the student data in the database matches the student data in the provided Excel file.
    Identifies the department and scheme for the extracted students and displays missing students accordingly.
    """
    try:
        cursor = connection.cursor(pymysql.cursors.DictCursor)

        # Read the Excel file to extract student data
        df = pd.read_excel(excel_file_path, sheet_name='Result Sheet')  # Adjust sheet name if needed
        extracted_students = set(zip(df['University Seat Number'], df['Student Name']))

        # Fetch all student details from the database
        query = "SELECT Usn, StudentName, dept, scheme FROM student"
        cursor.execute(query)
        db_students = {row['Usn']: (row['StudentName'], row['dept'], row['scheme']) for row in cursor.fetchall()}

        # Group extracted students by their department and scheme
        dept_scheme_map = {}
        missing_in_db = []
        missing_in_extracted = {}

        for usn, name in extracted_students:
            if usn in db_students:
                db_name, dept, scheme = db_students[usn]
                if name != db_name:
                    print(f"Name mismatch for USN: {usn}. DB: {db_name}, Extracted: {name}")

                # Group extracted students by dept and scheme
                key = (dept, scheme)
                if key not in dept_scheme_map:
                    dept_scheme_map[key] = set()
                dept_scheme_map[key].add((usn, name))
            else:
                missing_in_db.append((usn, name))

        # Compare students in each dept and scheme group
        for (dept, scheme), students in dept_scheme_map.items():
            query = "SELECT Usn, StudentName FROM student WHERE dept = %s AND scheme = %s"
            cursor.execute(query, (dept, scheme))
            db_group_students = set((row['Usn'], row['StudentName']) for row in cursor.fetchall())

            # Find missing students
            missing_in_group = db_group_students - students
            if missing_in_group:
                missing_in_extracted[(dept, scheme)] = missing_in_group

        # Report missing students
        if missing_in_db:
            print("\nStudents in extracted data but not found in the database:")
            for usn, name in missing_in_db:
                print(f"USN: {usn}, Name: {name}")

        if missing_in_extracted:
            for (dept, scheme), missing_students in missing_in_extracted.items():
                print(f"\nStudents in the database but not in extracted data (Dept: {dept}, Scheme: {scheme}):")
                for usn, name in missing_students:
                    print(f"USN: {usn}, Name: {name}")

        if not missing_in_db and not missing_in_extracted:
            print("Validation successful: All students match.")

    except Exception as e:
        print(f"Error during validation: {e}")


# Assuming these functions are defined somewhere in your code:
# create_tables(), extract_text_from_pdf(), extract_specific_sections(), extract_table_data(),
# calculate_grade_point(), insert_extracted_data(), insert_sgpa_data(), fetch_credits(), swap_columns(),
# create_result_counts_df(), get_top_toppers(), insert_topper_data(), validate_student_counts(), notify_faculty()

if __name__ == "__main__":

    department = sys.argv[1]
    scheme = sys.argv[2]
    sem = sys.argv[3]

    # Path to the folder containing PDFs
    folder_path = os.path.join("C:\\wamp64\\www\\project\\templates\\uploads", department, scheme, sem)
    semester_data = defaultdict(list)
    subject_details = {}
    usn_list = set()
    extracted_students = set()

    try:
        # Create database tables
        create_tables()

        # Iterate over the files in the folder and process PDF files
        num_items_in_folder = len([file_name for file_name in os.listdir(folder_path) if file_name.endswith('.pdf')])
        print(f"Number of items in folder: {num_items_in_folder}")
        for file_name in os.listdir(folder_path):
            if file_name.endswith('.pdf'):
                pdf_file_path = os.path.join(folder_path, file_name)
                extracted_text = extract_text_from_pdf(pdf_file_path)
                extracted_sections = extract_specific_sections(extracted_text)
                table_data = extract_table_data(extracted_text, extracted_sections)
                semester = extracted_sections.get("Semester", "Unknown Semester")
                semester_data[semester].extend(table_data)

        subject_results = defaultdict(lambda: Counter({'FCD': 0, 'FC': 0, 'SC': 0, 'Fail': 0}))

        # Creating DataFrame with the extracted data
        data = []
        for semester, data_points in semester_data.items():
            if data_points:
                for item in data_points:
                    if 'Result' not in item:
                        item['Result'] = 'Unknown'  # Default value for 'Result'
                    data.append(item)
            else:
                print(f"No data extracted for {semester}")

        if not data:
            print("No data to create DataFrame.")
        else:
            df = pd.DataFrame(data)
            df['Result'] = df['Result'].fillna('Unknown')
            print(df)

        total_counts = defaultdict(int)
        for semester, data in semester_data.items():
            insert_extracted_data(data)

            for row in data:
                usn = row['Usn']
                usn_list.add(usn)  # Collect unique USNs for validation
                subject_code = row['Subject Code']
                total_marks = row['Total Marks']
                row['Credit'] = credits
                row['Grade Point'] = calculate_grade_point(total_marks, credits)
                total_counts[subject_code] += 1
                if row.get('Result', '') == 'F':
                    subject_results[subject_code]['Fail'] += 1
                elif total_marks >= 70:
                    subject_results[subject_code]['FCD'] += 1
                elif total_marks >= 60:
                    subject_results[subject_code]['FC'] += 1
                elif total_marks >= 35:
                    subject_results[subject_code]['SC'] += 1
                else:
                    subject_results[subject_code]['Fail'] += 1

        # Creating the result counts DataFrame
        result_counts_df = create_result_counts_df(subject_results, total_counts)

        # Preparing DataFrame for all students
        df_list = []
        for semester, data in semester_data.items():
            usn_name_groups = defaultdict(list)
            for row in data:
                usn_name_groups[(row['Usn'], row['StudentName'])].append(row)
            for (usn, name), group in usn_name_groups.items():
                row_dict = {'Semester': semester, 'University Seat Number': usn, 'Student Name': name}
                total_credits = 0
                total_grade_points = 0
                total_marks = 0
                has_fail = any(row.get('Result', '') == 'F' for row in group)
                for row in group:
                    subject_code = row['Subject Code']
                    credits = fetch_credits(subject_code)
                    grade_points = row['Grade Point']
                    row['Credit'] = credits
                    row['C*GP'] = 0 if row.get('Result', '') == 'F' else credits * grade_points
                    total_credits += credits
                    total_grade_points += row['C*GP']
                    total_marks += row['Total Marks']
                sgpa = total_grade_points / total_credits if total_credits != 0 else 0
                rounded_sgpa = round(sgpa, 2)
                insert_sgpa_data(usn, semester, name, rounded_sgpa)

                for i, row in enumerate(group):
                    for key, value in row.items():
                        if key not in ('Semester', 'Usn', 'StudentName'):
                            row_dict[f'{key} {i+1}'] = value

                row_dict['Total C*GP'] = total_credits * rounded_sgpa
                row_dict['Total Credits'] = total_credits
                row_dict['SGPA'] = rounded_sgpa
                row_dict['Total Marks'] = total_marks

                percentage = rounded_sgpa * 10
                if has_fail:
                    result = 'Fail'
                elif percentage >= 70:
                    result = 'First Class with Distinction (FCD)'
                elif percentage >= 60:
                    result = 'First Class (FC)'
                elif percentage >= 35:
                    result = 'Second Class (SC)'
                else:
                    result = 'Fail'
                row_dict['%'] = f'{percentage:.2f}%'
                row_dict['Result'] = result

                df_list.append(row_dict)

        df = pd.DataFrame(df_list)
        columns_to_drop = [col for col in df.columns if 'AnnouncedDate' in col]
        df.drop(columns=columns_to_drop, inplace=True)

        rename_dict = {}
        for column in df.columns:
            if column[-1].isdigit():
                original_column_name = column.rsplit(' ', 1)[0]
                rename_dict[column] = original_column_name

        df.rename(columns=rename_dict, inplace=True)
        swap_columns(df, 'Result', 'Grade Point')
        swap_columns(df, 'Result', 'Credit')
        swap_columns(df, 'Result', 'C*GP')

        # Get top 3 toppers
        toppers_df = get_top_toppers(df)
        # Insert the top 3 toppers into the database
        insert_topper_data(toppers_df)

        # Define the base directory to save the Excel file
        base_dir = r"C:\\wamp64\\www\\project\\templates\\excel"

        # Create the path for department, scheme, and year
        save_dir = os.path.join(base_dir, department, scheme, sem)

        # Create directories if they do not exist
        os.makedirs(save_dir, exist_ok=True)

        # Define the file name
        file_name = f"{department}_{scheme}_{sem}SEM.xlsx"

        # Define the full path to save the Excel file
        excel_file_path = os.path.join(save_dir, file_name)


        # Check if the directory exists, if not, create it
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        # Create and write to the Excel file
        with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Result Sheet', index=False)
            worksheet = writer.sheets['Result Sheet']



            from openpyxl.styles import PatternFill
            # Yellow fill for fail students
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            # Get all column headers
            header_cells = [cell.value for cell in worksheet[1]]
            # Find the LAST occurrence of 'Result'
            result_col_idx = None
            for idx in range(len(header_cells)-1, -1, -1):  # Reverse loop
                if header_cells[idx] == 'Result':
                    result_col_idx = idx + 1  # Convert to 1-based index
                    break
            if result_col_idx is None:
                print("No 'Result' column found.")
            else:
                fail_count = 0
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    result_cell = row[result_col_idx - 1]  # 0-based index
                    if str(result_cell.value).strip() == "Fail":
                        fail_count += 1
                        for cell in row:
                            cell.fill = yellow_fill
                print(f"{fail_count} fail students highlighted succesfully!!")



            
            header_font = Font(bold=True)
            for cell in worksheet[1]:
                cell.font = header_font

            border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
            for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                for cell in row:
                    cell.border = border
            
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Write the result counts DataFrame to a new sheet
            result_counts_df.to_excel(writer, sheet_name='Results Count', index=False)

            # Write the top 3 toppers DataFrame to a new sheet
            toppers_df.to_excel(writer, sheet_name='Top 3 Toppers', index=False)

        # Validate student counts (ensure it's saved in the correct location)
        validate_student_counts(connection, excel_file_path)

        print(f"Excel file with results, counts, and top 3 toppers saved successfully at {excel_file_path}!")

        # Notify faculty members
        if os.path.exists(excel_file_path):
            notify_faculty(excel_file_path)
        else:
            print(f"Error: File not found at {excel_file_path}")

    finally:
        connection.close()
